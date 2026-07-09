import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime

def update_excel_archive(db, total_students_count: int, successful_count: int, failed_count: int):
    """
    Updates the master Excel archive 'Department_Analytics_Master.xlsx'.
    Maintains a 'Summary' sheet as Sheet 1, and appends/overwrites a date-based sheet
    for the current Sync All run.
    """
    filename = "Department_Analytics_Master.xlsx"
    sync_date_str = datetime.now().strftime("%d-%b-%Y")

    # 1. Load or initialize workbook
    if os.path.exists(filename):
        try:
            wb = openpyxl.load_workbook(filename)
        except Exception as e:
            print(f"Error loading existing workbook: {e}. Reinitializing.")
            wb = Workbook()
            summary_sheet = wb.active
            summary_sheet.title = "Summary"
            summary_sheet.append([
                "Sync Date",
                "Total Students",
                "Successful Students",
                "Failed Students",
                "Average Rating",
                "Average Problems Solved",
                "Average Easy",
                "Average Medium",
                "Average Hard"
            ])
    else:
        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        summary_sheet.append([
            "Sync Date",
            "Total Students",
            "Successful Students",
            "Failed Students",
            "Average Rating",
            "Average Problems Solved",
            "Average Easy",
            "Average Medium",
            "Average Hard"
        ])

    # Ensure Summary sheet is accessed correctly if workbook already existed
    if "Summary" in wb.sheetnames:
        summary_sheet = wb["Summary"]
    else:
        summary_sheet = wb.create_sheet("Summary", 0)
        summary_sheet.append([
            "Sync Date",
            "Total Students",
            "Successful Students",
            "Failed Students",
            "Average Rating",
            "Average Problems Solved",
            "Average Easy",
            "Average Medium",
            "Average Hard"
        ])

    # 2. Gather student standings details for the date sheet and summary averages
    from app.repositories.student_repository import get_all_students
    from app.repositories.profile_snapshot_repository import get_latest_snapshot
    from app.models.contest_result import ContestResult
    from app.models.contest import Contest

    students = get_all_students(db)
    date_rows = []

    ratings = []
    totals = []
    easies = []
    mediums = []
    hards = []

    for student in students:
        snapshot = get_latest_snapshot(db, student.id)

        # Retrieve rating & solves
        rating_val = snapshot.current_rating if (snapshot and snapshot.current_rating is not None) else None
        total_val = snapshot.total_solved if (snapshot and snapshot.total_solved is not None) else None
        easy_val = snapshot.easy_solved if (snapshot and snapshot.easy_solved is not None) else None
        medium_val = snapshot.medium_solved if (snapshot and snapshot.medium_solved is not None) else None
        hard_val = snapshot.hard_solved if (snapshot and snapshot.hard_solved is not None) else None

        if rating_val is not None:
            ratings.append(rating_val)
        if total_val is not None:
            totals.append(total_val)
        if easy_val is not None:
            easies.append(easy_val)
        if medium_val is not None:
            mediums.append(medium_val)
        if hard_val is not None:
            hards.append(hard_val)

        # Get Latest Weekly Contest Rank
        weekly_result = (
            db.query(ContestResult)
            .join(Contest, ContestResult.contest_id == Contest.id)
            .filter(ContestResult.student_id == student.id)
            .filter(Contest.contest_type.ilike("%weekly%"))
            .filter(~Contest.contest_type.ilike("%biweekly%"))
            .order_by(Contest.contest_date.desc())
            .first()
        )
        latest_weekly_rank = weekly_result.global_rank if (weekly_result and weekly_result.global_rank is not None) else "NA"

        # Get Latest Biweekly Contest Rank
        biweekly_result = (
            db.query(ContestResult)
            .join(Contest, ContestResult.contest_id == Contest.id)
            .filter(ContestResult.student_id == student.id)
            .filter(Contest.contest_type.ilike("%biweekly%"))
            .order_by(Contest.contest_date.desc())
            .first()
        )
        latest_biweekly_rank = biweekly_result.global_rank if (biweekly_result and biweekly_result.global_rank is not None) else "NA"

        date_rows.append([
            student.roll_no if student.roll_no else "NA",
            student.name if student.name else "NA",
            student.leetcode_username if student.leetcode_username else "NA",
            rating_val if rating_val is not None else "NA",
            total_val if total_val is not None else "NA",
            easy_val if easy_val is not None else "NA",
            medium_val if medium_val is not None else "NA",
            hard_val if hard_val is not None else "NA",
            latest_weekly_rank,
            latest_biweekly_rank
        ])

    # Calculate Averages (exclude NAs or use standard rounded integers)
    avg_rating = round(sum(ratings) / len(ratings)) if ratings else "NA"
    avg_total = round(sum(totals) / len(totals)) if totals else "NA"
    avg_easy = round(sum(easies) / len(easies)) if easies else "NA"
    avg_medium = round(sum(mediums) / len(mediums)) if mediums else "NA"
    avg_hard = round(sum(hards) / len(hards)) if hards else "NA"

    # 3. Create or Overwrite Date-based Sheet
    if sync_date_str in wb.sheetnames:
        wb.remove(wb[sync_date_str])

    date_sheet = wb.create_sheet(sync_date_str)
    date_headers = [
        "Roll No",
        "Student Name",
        "LeetCode Username",
        "Contest Rating",
        "Total Solved",
        "Easy Solved",
        "Medium Solved",
        "Hard Solved",
        "Latest Weekly Contest Rank",
        "Latest Biweekly Contest Rank"
    ]
    date_sheet.append(date_headers)
    for row in date_rows:
        date_sheet.append(row)

    # 4. Insert or Update Summary Sheet Log Row
    summary_row = [
        sync_date_str,
        total_students_count,
        successful_count,
        failed_count,
        avg_rating,
        avg_total,
        avg_easy,
        avg_medium,
        avg_hard
    ]

    # Find if row for this date already exists in the Summary table
    found_row_idx = None
    for r_idx in range(2, summary_sheet.max_row + 1):
        cell_val = summary_sheet.cell(row=r_idx, column=1).value
        if cell_val == sync_date_str:
            found_row_idx = r_idx
            break

    if found_row_idx is not None:
        # Overwrite existing row
        for c_idx, val in enumerate(summary_row, 1):
            summary_sheet.cell(row=found_row_idx, column=c_idx, value=val)
    else:
        # Append new row
        summary_sheet.append(summary_row)

    # 5. Save master file
    wb.save(filename)
    print(f"Excel Master archive successfully updated: '{sync_date_str}' sheet saved.")
